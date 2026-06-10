# -*- coding: utf-8 -*-
"""
주간회의자료(주식 시황회의) 자동 생성기 — 원본 양식 보존(템플릿 방식).

기존 수기 .xlsx 를 **템플릿으로 로드**해 서식·수식·병합을 그대로 두고
데이터 셀만 최신값으로 덮어쓴다. 따라서 원본과 양식이 100% 동일.

자동 갱신 영역:
  · 주간 주식시장 변동 — 우측 raw 종가표(P8:X11)만 갱신 → 좌측은 수식이 자동 반영
    (한 KOSPI/KOSDAQ, 미 S&P/나스닥, 유럽 EuroStoxx50, 일 Nikkei, 중 상해, MSCI ACWI/EM)
  · 주요 Factor 점검 — macro-data.js(지표·해석 commentary) + kr_flows.json(수급) 실데이터
    기업이익/경제지표/수급/주요이벤트/주요보고서 = 지표값 + 해석을 함께 기재
  · 주식 손익 — portfolio-data ytdMetrics(미실현/실현) 이식, 1~5월/6월/계 분해
  · 국가·섹터 비중 — region / sector_exposure 룩스루
  · (추가) □ 시장 센티먼트 — CNN Fear&Greed + AAII Bull/Bear (원본 하단에 덧붙임)

사용법:
  python weekly/weekly_report.py [--date YYYY-MM-DD] [--out PATH]
                                 [--template PATH] [--today YYYY-MM-DD] [--no-sentiment]
"""
import sys, os, re, io, json, argparse, glob
from datetime import datetime, date, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
ONEDRIVE_BASE = os.path.join(os.path.expanduser("~"), "OneDrive", "삼성화재 주간 발표자료")
F_NAME = "맑은 고딕"


# ════════════════════════ 데이터 로드 ════════════════════════
def _parse(path, var_re, whole=True):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    m = re.search(var_re, text, re.S)
    if not m:
        raise ValueError("assign not found: " + path)
    return json.loads(m.group(1))


def load_benchmarks():
    return _parse(os.path.join(REPO, "benchmarks.js"),
                  r"window\.BENCHMARKS\s*=\s*(\{.*\})\s*;?\s*$")


def load_macro():
    try:
        return _parse(os.path.join(REPO, "macro-data.js"),
                      r"=\s*(\{.*\})\s*;?\s*$")
    except Exception as e:
        sys.stderr.write("macro load fail: %s\n" % e)
        return {}


def load_kr_flows():
    try:
        with open(os.path.join(REPO, "kr_flows.json"), "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def load_daily():
    try:
        return _parse(os.path.join(REPO, "daily-data.js"),
                      r"window\.DAILY\s*=\s*(\{.*\})\s*;?\s*$")
    except Exception:
        return None


def load_portfolio():
    import subprocess
    plain = os.path.join(REPO, "portfolio-data.plain.js")
    if not os.path.exists(plain):
        subprocess.run([sys.executable, "encrypt_data.py", "decrypt"],
                       cwd=REPO, capture_output=True)
    with open(plain, "r", encoding="utf-8") as f:
        txt = f.read()
    # PORTFOLIO_DATA = { ... };  뒤에 보조 할당이 이어지므로 raw_decode 로 첫 객체만 파싱.
    m = re.search(r"PORTFOLIO_DATA\s*=\s*", txt)
    start = txt.index("{", m.end())
    obj, _ = json.JSONDecoder().raw_decode(txt, start)
    return obj


def get_sentiment(enabled=True):
    if not enabled:
        return {"cnn_fng": None, "aaii": None}
    try:
        sys.path.insert(0, HERE)
        import fetch_sentiment
        return fetch_sentiment.collect()
    except Exception as e:
        sys.stderr.write("sentiment fail: %s\n" % e)
        return {"cnn_fng": None, "aaii": None}


# ════════════════════════ 시장 종가(raw) ════════════════════════
# 우측표 컬럼 → benchmarks ticker. T(유럽)=EuroStoxx50 은 yfinance 로 별도.
COL_TICKER = {"P": "^KS11", "Q": "^KQ11", "R": "^GSPC", "S": "^IXIC",
              "U": "^N225", "V": "000001.SS", "W": "ACWI", "X": "EEM"}


def _hist_level_at(idx, target):
    """history 에서 target(date) 이하 마지막 종가."""
    h = idx.get("history") or {}
    dates, vals = h.get("dates") or [], h.get("values") or []
    lvl = None
    for d, v in zip(dates, vals):
        try:
            dd = datetime.strptime(d, "%Y-%m-%d").date()
        except Exception:
            continue
        if dd <= target:
            lvl = v
    return lvl


def bench_levels(bench, as_of_d):
    """각 컬럼별 (연말, 현재, MTD기준, 1W기준) raw 종가."""
    by_t = {x.get("ticker"): x for x in (bench.get("indices") or [])}
    out = {}
    week_ago = as_of_d - timedelta(days=7)
    for col, tk in COL_TICKER.items():
        x = by_t.get(tk)
        if not x:
            out[col] = (None, None, None, None)
            continue
        out[col] = (x.get("baseline"), x.get("current"),
                    x.get("mtd_baseline"), _hist_level_at(x, week_ago))
    return out


def euro_stoxx_levels(as_of_d):
    """^STOXX50E raw 종가 (연말2025, 현재, 5월말, 1주전)."""
    try:
        import yfinance as yf
    except ImportError:
        return (None, None, None, None)
    try:
        h = yf.Ticker("^STOXX50E").history(
            start="2025-12-15", end=(as_of_d + timedelta(days=2)).isoformat(),
            auto_adjust=False)[["Close"]].dropna()
        if h.empty:
            return (None, None, None, None)
        def last_on_or_before(d):
            sub = h[h.index.date <= d]
            return round(float(sub.iloc[-1]["Close"]), 2) if len(sub) else None
        ye = last_on_or_before(date(2025, 12, 31))
        cur = round(float(h.iloc[-1]["Close"]), 2)
        may = last_on_or_before(date(as_of_d.year, as_of_d.month, 1) - timedelta(days=1))
        wk = last_on_or_before(as_of_d - timedelta(days=7))
        return (ye, cur, may, wk)
    except Exception as e:
        sys.stderr.write("euro stoxx fail: %s\n" % e)
        return (None, None, None, None)


# ════════════════════════ 포트폴리오 손익/비중 ════════════════════════
SECTOR_KO = {"IT": "IT", "Financials": "금융", "Healthcare": "헬스케어",
             "Industrials": "산업재", "Cons Disc": "경기소비", "Communication": "커뮤니케이션",
             "Cons Staples": "필수소비", "Energy": "에너지", "Materials": "소재",
             "Utilities": "유틸리티", "Real Estate": "부동산"}
KO_SECTOR = {v: k for k, v in SECTOR_KO.items()}
REGION_BUCKET = {"한국": "한국", "미국": "미국", "유럽": "유럽", "이머징": "이머징", "글로벌": "기타"}


def ytd_metrics(h):
    start = h.get("start_2026")
    if start is None:
        start = h.get("book") or 0
    buy = h.get("ytd_buy") or 0
    sell = h.get("ytd_sell") or 0
    mkt = h.get("mkt") or 0
    cost = start + buy
    bb = h.get("book_basis")
    if bb is not None:
        cost_rem = bb
        cost_sold = cost - cost_rem
    elif sell > 0 and (mkt + sell) > 0:
        cost_sold = cost * sell / (mkt + sell)
        cost_rem = cost - cost_sold
    else:
        cost_sold = 0
        cost_rem = cost
    return {"realized": sell - cost_sold, "unrealized": mkt - cost_rem}


def snapshot_holdings(pdata, cutoff):
    hist = pdata.get("historical") or {}
    hm = hist.get("holdings_mkt") or {}
    dates = hist.get("dates") or []
    if not dates:
        return None
    idx = None
    for i, d in enumerate(dates):
        if d <= cutoff:
            idx = i
    if idx is None:
        return None
    trades = pdata.get("trades") or []
    out = []
    for h in pdata.get("holdings") or []:
        ser = hm.get(h.get("name"))
        mkt_at = ser[idx] if (ser and idx < len(ser)) else (h.get("mkt") or 0)
        isin = h.get("isin") or ""
        buy = sell = 0.0
        for t in trades:
            if not isin or t.get("isin") != isin or (t.get("date") or "") > cutoff:
                continue
            if t.get("action") == "매입":
                buy += t.get("amount") or 0
            elif t.get("action") == "매도":
                sell += t.get("amount") or 0
        start = h.get("start_2026")
        if start is None:
            start = h.get("book") or 0
        fs, fc = h.get("ytd_sell") or 0, h.get("cost_sold") or 0
        cost_sold_at = fc * (sell / fs) if (fs > 0 and sell > 0) else 0
        out.append({"mkt": mkt_at, "start_2026": start, "ytd_buy": buy,
                    "ytd_sell": sell, "cost_sold": cost_sold_at,
                    "book_basis": (start + buy) - cost_sold_at})
    return out


def agg(holdings):
    r = u = 0.0
    for h in holdings:
        m = ytd_metrics(h)
        r += m["realized"]
        u += m["unrealized"]
    return {"realized": r, "unrealized": u, "pnl": r + u}


def build_pnl(pdata):
    ytd = agg(pdata.get("holdings") or [])
    snap = snapshot_holdings(pdata, "2026-05-31")
    may = agg(snap) if snap else None
    jun = ({"realized": ytd["realized"] - may["realized"],
            "unrealized": ytd["unrealized"] - may["unrealized"],
            "pnl": ytd["pnl"] - may["pnl"]} if may else None)
    return {"ytd": ytd, "to_may": may, "jun": jun}


def build_weights(pdata):
    H = pdata.get("holdings") or []
    total = sum(h.get("mkt") or 0 for h in H) or 1
    country = {}
    for h in H:
        b = REGION_BUCKET.get(h.get("region"), "기타")
        country[b] = country.get(b, 0) + (h.get("mkt") or 0) / total
    sec = {}
    for h in H:
        mkt = h.get("mkt") or 0
        for s, e in (h.get("sector_exposure") or {}).items():
            sec[s] = sec.get(s, 0) + (mkt / total) * e
    return country, {SECTOR_KO.get(k, k): v for k, v in sec.items()}


# ════════════════════════ Factor 점검 내용(실데이터+해석) ════════════════════════
def _ind(macro, key, field="current"):
    try:
        return macro["indicators"][key][field]
    except Exception:
        return None


def _fnum(v, suf="", plus=False):
    if v is None:
        return "-"
    s = ("%+.1f" if plus else "%.1f") % v
    return s + suf


def build_factor_lines(macro, kr):
    comm = (macro or {}).get("commentary") or {}
    earn = ((macro or {}).get("earnings") or {}).get("countries") or {}
    us, kre = earn.get("US") or {}, earn.get("KR") or {}
    pmi = None
    mm = re.search(r"PMI\s*([\d.]+)", comm.get("macro", "") or "")
    if mm:
        pmi = mm.group(1)

    기업이익 = [
        " - 미국 ERR(이익수정비율) %s(%s), Fwd EPS수정 1M %s%% / 3M %s%%" % (
            _g(us, "err"), _g(us, "err_label"),
            _fnum(us.get("rev30"), plus=True), _fnum(us.get("rev90"), plus=True)),
        " - 한국 ERR %s(%s), Fwd EPS수정 1M %s%%" % (
            _g(kre, "err"), _g(kre, "err_label"), _fnum(kre.get("rev30"), plus=True)),
        " - '26 EPS성장 전망: 미국 %s%%, 한국 %s%%" % (
            _fnum(us.get("growth_cy")), _fnum(kre.get("growth_cy"))),
        " ※ %s" % comm.get("earnings", ""),
    ]
    경제지표 = [
        " - 미국 CPI(YoY) %s%% / 근원 CPI %s%%" % (
            _fnum(_ind(macro, "cpi_yoy")), _fnum(_ind(macro, "core_cpi_yoy"))),
        " - 미국 실업률 %s%% / 비농업고용 %s천명(전월비)" % (
            _fnum(_ind(macro, "unemployment")), _fnum(_ind(macro, "payrolls"))),
        " - 미 기준금리 %s%%%s" % (
            _fnum(_ind(macro, "fed_funds")),
            (" / ISM 제조업 PMI %s" % pmi) if pmi else ""),
        " - 밸류: S&P500 12M Fwd PER %s배 / KOSPI %s배" % (
            _fnum(_ind(macro, "spx_fwd_pe")), _fnum(_ind(macro, "kospi_fwd_pe"))),
        " ※ 해석(매크로): %s" % comm.get("macro", ""),
        " ※ 해석(밸류): %s" % comm.get("valuation", ""),
    ]
    mtd = (kr or {}).get("mtd") or {}
    ytd = (kr or {}).get("ytd_total") or {}
    unit = (kr or {}).get("unit", "조원")
    수급 = [
        " - (국내) 외국인 6월 %s%s(YTD %s%s) — 매도세 지속" % (
            _fnum(mtd.get("foreign"), plus=True), unit,
            _fnum(ytd.get("foreign"), plus=True), unit),
        " - 개인 %s%s, 기관 %s%s (6월 %s거래일)" % (
            _fnum(mtd.get("retail"), plus=True), unit,
            _fnum(mtd.get("inst"), plus=True), unit, mtd.get("days", "-")),
        " ※ 해석: %s" % comm.get("flows", ""),
    ]
    events = ((macro or {}).get("monthly_factors") or {}).get("events") or []
    주요이벤트 = []
    for e in events[:5]:
        imp = (e.get("impact") or "").split(".")[0][:60]
        주요이벤트.append("  · %s %s — %s" % (e.get("date", ""), e.get("title", ""), imp))
    if not 주요이벤트:
        주요이벤트 = ["  · 주요 일정 기재"]
    out = ((macro or {}).get("outlook") or {})
    주요보고서 = [
        " - (단기) %s: %s" % (_g(out.get("short", {}), "bias"), _g(out.get("short", {}), "text")),
        " - (중기) %s: %s" % (_g(out.get("mid", {}), "bias"), _g(out.get("mid", {}), "text")),
        " - (장기) %s: %s" % (_g(out.get("long", {}), "bias"), _g(out.get("long", {}), "text")),
        " ※ 종합: %s" % comm.get("overall", ""),
    ]
    return {"기업이익": (기업이익, 13), "경제지표": (경제지표, 17),
            "수급": (수급, 23), "주요이벤트": (주요이벤트, 28), "주요보고서": (주요보고서, 33)}


def _g(d, k):
    v = (d or {}).get(k)
    return "-" if v is None else v


def _std_key(k):
    return _FACTOR_LABELS.get(_norm(k), k)


def sentiment_fallback_lines(sentiment, bench, macro):
    """enrich.json 이 없을 때 센티먼트 섹션 자동 생성."""
    cnn = (sentiment or {}).get("cnn_fng") or {}
    aaii = (sentiment or {}).get("aaii") or {}
    vix = next((x for x in (bench.get("indices") or []) if x.get("ticker") == "^VIX"), {})
    out = []
    if cnn.get("score") is not None:
        out.append(" - CNN Fear&Greed %.0f(%s) — 1주전 %s·1개월전 %s·1년전 %s" % (
            cnn["score"], _rate_ko(cnn.get("rating")), _n(cnn.get("prev_1week")),
            _n(cnn.get("prev_1month")), _n(cnn.get("prev_1year"))))
    if aaii.get("bullish") is not None:
        out.append(" - AAII 강세 %.0f%%·약세 %.0f%% (Bull-Bear %+.0f%%p)" % (
            aaii["bullish"], aaii["bearish"], aaii.get("spread") or 0))
    if vix:
        out.append(" - VIX %.1f (일간 %+.1f%%)" % (vix.get("current") or 0, vix.get("daily_pct") or 0))
    return out or [" - 센티먼트 지표 기재"]


def resolve_factor_content(macro, kr, sentiment, bench):
    """Factor 콘텐츠 {표준키: [라인]}. enrich.json(웹 보강) 우선, 없으면 macro 자동."""
    enrich = load_enrich()
    if enrich:
        return {_std_key(k): v for k, v in enrich.items()}
    fb = build_factor_lines(macro, kr)  # {라벨:(lines,row)}
    out = {_std_key(k): v[0] for k, v in fb.items()}
    out.setdefault("센티먼트", sentiment_fallback_lines(sentiment, bench, macro))
    return out


# ════════════════════════ 템플릿 ════════════════════════
def find_template(explicit=None):
    if explicit:
        return explicit
    # 사용자가 손본 양식 우선순위: 주간회의자료_.xlsx > 주간회의자료.xlsx (auto 출력 제외).
    best = None
    for name in ("주간회의자료_.xlsx", "주간회의자료.xlsx"):
        cands = []
        for p in glob.glob(os.path.join(ONEDRIVE_BASE, "*", name)):
            folder = os.path.basename(os.path.dirname(p))
            if re.fullmatch(r"\d{8}", folder):
                cands.append((folder, p))
        if cands:
            cands.sort()
            best = cands[-1][1]
            break
    if not best:
        raise FileNotFoundError("템플릿(주간회의자료*.xlsx)을 찾을 수 없음: " + ONEDRIVE_BASE)
    return best


def load_enrich():
    """weekly/enrich.json (웹 리서치 보강 콘텐츠) 로드. 없으면 None."""
    p = os.path.join(HERE, "enrich.json")
    if not os.path.exists(p):
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            d = json.load(f)
        return {k: v for k, v in d.items() if not k.startswith("_")}
    except Exception:
        return None


def load_current_override():
    """weekly/current_override.json (yfinance 지연 시 현재가 수동 보정). 없으면 None."""
    p = os.path.join(HERE, "current_override.json")
    if not os.path.exists(p):
        return None
    try:
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


# Factor 섹션 라벨(정규화) → 표준 키. 사용자가 띄어쓰기를 바꿔도 매칭.
_FACTOR_LABELS = {
    "기업이익": "기업이익", "경제지표": "경제지표", "수급": "수급",
    "주요이벤트": "주요 이벤트", "주요이벤트점검": "주요 이벤트",
    "센티먼트": "센티먼트", "센티멘트": "센티먼트",
    "주요보고서": "주요 보고서",
}


def _norm(s):
    return re.sub(r"\s+", "", str(s or ""))


def detect_factor_rows(ws, max_row=60):
    """col B 라벨을 스캔해 각 Factor 섹션의 (시작행, 끝행, 표준키)."""
    hits = []  # (row, key)
    for r in range(11, max_row):
        b = ws["B%d" % r].value
        key = _FACTOR_LABELS.get(_norm(b))
        if key:
            hits.append((r, key))
    out = {}
    for i, (r, key) in enumerate(hits):
        end = (hits[i + 1][0] - 1) if i + 1 < len(hits) else r + 5
        # 다음 □ 섹션 전까지로 제한
        for rr in range(r + 1, end + 1):
            bv = ws["B%d" % rr].value
            if bv and str(bv).strip().startswith("□"):
                end = rr - 1
                break
        out[key] = (r, end)
    return out


def detect_pnl_rows(ws, max_row=60):
    """FVPL 행과 평가/매각·배당/계 행 + 1~5월/6월/계 컬럼 탐지."""
    fvpl = None
    for r in range(38, max_row):
        if _norm(ws["B%d" % r].value) == "FVPL":
            fvpl = r
            break
    if fvpl is None:
        return None
    rows = {}
    for r in range(fvpl, fvpl + 4):
        c = _norm(ws["C%d" % r].value)
        if c in ("평가", "매각/평가", "평가/매각"):
            rows["unrealized"] = r
        elif c in ("매각/배당", "이자", "매각배당"):
            rows["realized"] = r
        elif c == "계":
            rows["pnl"] = r
    return rows or None


def detect_weight_rows(ws, max_row=60):
    """국가(col H→I) / 섹터(col J→K) 라벨 위치 탐지."""
    country, sector = {}, {}
    for r in range(40, max_row):
        h = (ws["H%d" % r].value or "")
        if h in ("한국", "미국", "유럽", "이머징", "기타"):
            country[h] = r
        j = ws["J%d" % r].value
        if j:
            sector[str(j).strip()] = r
    return country, sector


def stock_sheet(wb):
    for ws in wb.worksheets:
        if ws["D4"].value == "KOSPI":
            return ws
    return wb.worksheets[-1]


def _set(ws, coord, value):
    ws[coord] = value  # 기존 서식/수식포맷 유지


# ════════════════════════ 빌드 ════════════════════════
def build(target, bench, macro, kr, pdata, daily, sentiment, template):
    # 템플릿이 Excel/OneDrive 로 잠겨 있을 수 있으므로 임시본을 만들어 로드.
    import tempfile, shutil
    tmpl_copy = os.path.join(tempfile.gettempdir(), "_wk_template.xlsx")
    try:
        shutil.copy2(template, tmpl_copy)
        src = tmpl_copy
    except Exception:
        src = template
    wb = openpyxl.load_workbook(src)  # 서식 보존
    ws = stock_sheet(wb)
    ws.title = target.strftime("%Y%m%d")
    wk = (target.day - 1) // 7 + 1

    # 제목
    _set(ws, "A1", "주식 시황회의(%d월 %d주차)" % (target.month, wk))

    # ── 시장 종가표 ─ 우측 raw(P8:X11) + 좌측 표시값(rows 8·9) 동시 갱신 ──
    as_of_d = datetime.strptime(bench.get("as_of"), "%Y-%m-%d").date()
    lv = bench_levels(bench, as_of_d)
    lv["T"] = euro_stoxx_levels(as_of_d)  # 유럽 EuroStoxx50
    for col, (ye, cur, mb, wkb) in lv.items():
        for row, val in ((8, ye), (9, cur), (10, mb), (11, wkb)):
            if val is not None:
                _set(ws, "%s%d" % (col, row), round(val, 2))
    # yfinance 지연 등으로 현재가가 stale 할 때, 수동 오버라이드(현재=row9)를 적용.
    ovr = load_current_override()
    if ovr and ovr.get("as_of") == target.strftime("%Y-%m-%d"):
        for col, val in (ovr.get("levels") or {}).items():
            if val is not None:
                _set(ws, "%s9" % col, round(float(val), 2))
    # 좌측 표(D~L)의 연말·현재 표시값은 수식이 아닌 하드값이므로 우측과 동일하게 복사.
    LR = [("D", "P"), ("E", "Q"), ("F", "R"), ("G", "S"), ("H", "T"),
          ("I", "U"), ("J", "V"), ("K", "W"), ("L", "X")]
    for lc, rc in LR:
        for row in (8, 9):
            v = ws["%s%d" % (rc, row)].value
            if v is not None:
                _set(ws, "%s%d" % (lc, row), v)
    # 날짜 라벨
    _set(ws, "N9", datetime.combine(target, datetime.min.time()))
    _set(ws, "B9", datetime.combine(target, datetime.min.time()))
    prev_month_end = date(as_of_d.year, as_of_d.month, 1) - timedelta(days=1)
    _set(ws, "N10", datetime.combine(prev_month_end, datetime.min.time()))
    _set(ws, "N11", datetime.combine(as_of_d - timedelta(days=7), datetime.min.time()))

    # ── 주요 Factor 점검 ─ 섹션 행을 동적 탐지, 콘텐츠는 enrich.json(웹 보강) 우선 ──
    content = resolve_factor_content(macro, kr, sentiment, bench)
    sections = detect_factor_rows(ws)
    for key, (start, end) in sections.items():
        lines = content.get(key) or []
        for rr in range(start, end + 1):  # 기존 내용 클리어
            ws["D%d" % rr] = None
        for i, ln in enumerate(lines):
            if start + i <= end:
                _set(ws, "D%d" % (start + i), ln)

    # ── 주식 손익 (FVPL 평가/매각·배당/계 × 1~5월/6월/계) ─ 동적 행 ──
    pnl = build_pnl(pdata)
    prows = detect_pnl_rows(ws)
    if prows:
        for key, row in prows.items():
            may = pnl["to_may"][key] if pnl["to_may"] else None
            jun = pnl["jun"][key] if pnl["jun"] else None
            _set(ws, "D%d" % row, round(may, 1) if may is not None else None)
            _set(ws, "E%d" % row, round(jun, 1) if jun is not None else None)
            _set(ws, "F%d" % row, round(pnl["ytd"][key], 1))

    # ── 국가·섹터 비중 ─ 동적 행 ──
    country, sec = build_weights(pdata)
    crows, srows = detect_weight_rows(ws)
    for ko, row in crows.items():
        if ko in country:
            _set(ws, "I%d" % row, round(country[ko], 4))
    for label, row in srows.items():
        if label in sec:
            _set(ws, "K%d" % row, round(sec[label], 4))

    # 좌측 표가 우측 raw 종가를 참조하는 수식이므로, 열 때 강제 재계산.
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    return wb


def add_sentiment_block(ws, sentiment, bench, macro, start_row):
    bold = Font(name=F_NAME, size=11, bold=True)
    reg = Font(name=F_NAME, size=10)
    fill = PatternFill("solid", fgColor="FFF2CC")
    r = start_row
    ws["B%d" % r] = "□ 시장 센티먼트"
    ws["B%d" % r].font = Font(name=F_NAME, size=12, bold=True)
    r += 1
    cnn = (sentiment or {}).get("cnn_fng")
    aaii = (sentiment or {}).get("aaii")
    if cnn and cnn.get("score") is not None:
        ws["B%d" % r] = "CNN Fear & Greed"; ws["B%d" % r].font = bold
        ws["B%d" % r].fill = fill
        ws["D%d" % r] = "%.0f (%s)" % (cnn["score"], _rate_ko(cnn.get("rating")))
        ws["D%d" % r].font = bold
        ws["F%d" % r] = "1주전 %s · 1개월전 %s · 1년전 %s" % (
            _n(cnn.get("prev_1week")), _n(cnn.get("prev_1month")), _n(cnn.get("prev_1year")))
        ws["F%d" % r].font = reg
        r += 1
        comps = cnn.get("components") or []
        txt = " · ".join("%s %.0f" % (c["label"], c["score"]) for c in comps)
        ws["B%d" % r] = " 구성"; ws["B%d" % r].font = reg
        ws["D%d" % r] = txt; ws["D%d" % r].font = Font(name=F_NAME, size=9)
        r += 1
    if aaii and aaii.get("bullish") is not None:
        ws["B%d" % r] = "AAII 투자심리"; ws["B%d" % r].font = bold
        ws["B%d" % r].fill = fill
        ws["D%d" % r] = "강세 %.0f%% / 중립 %.0f%% / 약세 %.0f%%" % (
            aaii["bullish"], aaii["neutral"], aaii["bearish"])
        ws["D%d" % r].font = reg
        sp = aaii.get("spread")
        ws["F%d" % r] = ("Bull-Bear %+.0f%%p (역사평균 강세 %.1f%%)" %
                         (sp, aaii.get("bull_hist_avg") or 0)) if sp is not None else ""
        ws["F%d" % r].font = reg
        r += 1
    vix = next((x for x in (bench.get("indices") or []) if x.get("ticker") == "^VIX"), None)
    pc = None
    mc = re.search(r"풋/콜[^0-9]*([\d.]+)", ((macro or {}).get("commentary") or {}).get("sentiment", "") or "")
    if mc:
        pc = mc.group(1)
    if vix:
        ws["B%d" % r] = "VIX / 풋콜"; ws["B%d" % r].font = bold
        ws["B%d" % r].fill = fill
        ws["D%d" % r] = "VIX %.1f (일간 %+.1f%%)%s" % (
            vix.get("current") or 0, vix.get("daily_pct") or 0,
            (" · CBOE 풋/콜 %s" % pc) if pc else "")
        ws["D%d" % r].font = reg
        r += 1


def _rate_ko(rating):
    return {"extreme fear": "극단적 공포", "fear": "공포", "neutral": "중립",
            "greed": "탐욕", "extreme greed": "극단적 탐욕"}.get((rating or "").lower(), rating or "")


def _n(v):
    return ("%.0f" % v) if isinstance(v, (int, float)) else "-"


# ════════════════════════ main ════════════════════════
def next_monday(today):
    return today if today.weekday() == 0 else today + timedelta(days=(7 - today.weekday()))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date", default=None)
    ap.add_argument("--out", default=None)
    ap.add_argument("--template", default=None)
    ap.add_argument("--today", default=None)
    ap.add_argument("--no-sentiment", action="store_true")
    args = ap.parse_args()

    today = datetime.strptime(args.today, "%Y-%m-%d").date() if args.today else date.today()
    target = (datetime.strptime(args.date, "%Y-%m-%d").date() if args.date
              else next_monday(today))

    bench = load_benchmarks()
    macro = load_macro()
    kr = load_kr_flows()
    pdata = load_portfolio()
    daily = load_daily()
    sentiment = get_sentiment(not args.no_sentiment)
    template = find_template(args.template)

    wb = build(target, bench, macro, kr, pdata, daily, sentiment, template)

    out = args.out
    if not out:
        folder = os.path.join(ONEDRIVE_BASE, target.strftime("%Y%m%d"))
        os.makedirs(folder, exist_ok=True)
        out = os.path.join(folder, "주간회의자료_auto.xlsx")
    try:
        wb.save(out)
    except PermissionError:
        base, ext = os.path.splitext(out)
        out = base + "_" + datetime.now().strftime("%H%M%S") + ext
        wb.save(out)
    print("saved:", out)
    print("  template:", os.path.basename(os.path.dirname(template)),
          "| sentiment cnn:", bool(sentiment.get("cnn_fng")),
          "aaii:", bool(sentiment.get("aaii")))


if __name__ == "__main__":
    main()
